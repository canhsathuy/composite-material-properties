from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# Duong dan vao thu muc
path = "stress_data.xlsx"
# Tạo workbook mới
wb = load_workbook(path)
ws = wb["mesh size 10mm"]

# Đọc dữ liệu từ file Fortran xuất ra
with open("data.txt", "r") as file:
    for i, line in enumerate(file, start=1):  # Bắt đầu từ dòng 2 (sau tiêu đề)
        x, sigma_cap = map(float, line.split())
        ws.cell(row=i+3, column=14, value=x*1000)
        ws.cell(row=i+3, column=15, value=sigma_cap/1000000)
    font_style = Font(name = "Times New Roman", size = 12)
    for row in range(4,105):
        for col in range(14,16):
            cell = ws.cell(row = row, column = col)
            cell.font = font_style

# Lưu file Excel
wb.save(path)
print("Dữ liệu đã được ghi vào file 'stress_data.xlsx'")
